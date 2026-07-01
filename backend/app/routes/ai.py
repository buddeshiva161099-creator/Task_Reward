"""
AI Routes - Endpoints for AI Workforce Intelligence.
"""
from fastapi import APIRouter, Depends, Query, HTTPException, Body
from fastapi.responses import StreamingResponse, HTMLResponse
from app.auth.dependencies import get_current_user, require_management_team
from app.auth.tenant_scope import get_active_business_unit_id
from app.models.user import User, UserRole
from app.services import ai_service
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

router = APIRouter(prefix="/ai", tags=["AI Intelligence"])


@router.get("/dashboard-summary")
async def get_dashboard_summary(
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Get customized role-based AI dashboard summary, recommendations, and operational alerts."""
    return await ai_service.get_dashboard_intelligence_summary(current_user, business_unit_id=active_bu_id)


@router.get("/task-intelligence")
async def get_task_intelligence(
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Retrieve AI-driven predictions on task completion risk, overloaded assignees, and allocation suggestions."""
    scope = await ai_service.get_employee_ids_in_scope(current_user)
    return await ai_service.run_task_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)


@router.get("/performance-intelligence")
async def get_performance_intelligence(
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Retrieve AI-powered employees productivity index, work consistency rating, and burnout warning states."""
    scope = await ai_service.get_employee_ids_in_scope(current_user)
    return await ai_service.run_performance_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)


@router.get("/payroll-intelligence")
async def get_payroll_intelligence(
    current_user: User = Depends(require_management_team),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Scan and retrieve payroll outliers, overtime spikes, and deduction variance alerts (Management only)."""
    scope = await ai_service.get_employee_ids_in_scope(current_user)
    return await ai_service.run_payroll_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)


@router.get("/attendance-intelligence")
async def get_attendance_intelligence(
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Evaluate attendance log logs to predict absenteeism risks and late check-in login trends."""
    scope = await ai_service.get_employee_ids_in_scope(current_user)
    return await ai_service.run_attendance_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)


class AssistantPayload(BaseModel):
    message: str
    history: List[Dict[str, str]] = []


@router.post("/assistant")
async def ask_ai_assistant(
    payload: AssistantPayload,
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """In-app chat copilot accepting natural queries and returning permission-safe data insights."""
    if not payload.message.strip():
        raise HTTPException(status_code=400, detail="Message content cannot be empty.")
    return await ai_service.run_ai_copilot_assistant(
        user_message=payload.message,
        current_user=current_user,
        business_unit_id=active_bu_id,
        history=payload.history
    )


@router.get("/reports/export")
async def export_ai_report(
    report_type: str = Query("productivity", regex="^(productivity|payroll|team_performance|attendance|executive)$"),
    report_format: str = Query("excel", regex="^(excel|html)$"),
    current_user: User = Depends(get_current_user),
    active_bu_id = Depends(get_active_business_unit_id),
):
    """Generates and downloads reports (Excel sheet or styled print-ready HTML page) enriched with AI insights."""
    scope = await ai_service.get_employee_ids_in_scope(current_user)

    # Restrict payroll to Admin/HR/Manager
    if report_type == "payroll" and current_user.role not in [UserRole.ADMIN, UserRole.HR_MANAGER, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="Unauthorized role access to payroll reports.")

    # 1. Fetch AI operational summary context
    task_intel = await ai_service.run_task_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)
    perf_intel = await ai_service.run_performance_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)
    attendance_intel = await ai_service.run_attendance_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)

    payroll_intel = {"alerts": []}
    if current_user.role in [UserRole.ADMIN, UserRole.HR_MANAGER]:
        payroll_intel = await ai_service.run_payroll_analysis(scope, current_user.tenant_id, business_unit_id=active_bu_id)

    # 2. Build Content based on format
    if report_format == "excel":
        output = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AI Report Overview"
        
        # Styles
        title_font = Font(name='Arial', size=16, bold=True, color='3F51B5')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        bold_font = Font(name='Arial', size=11, bold=True)
        regular_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color='3F51B5', end_color='3F51B5', fill_type='solid')
        accent_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
        alert_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Title Block
        ws.merge_cells('A1:E1')
        ws['A1'] = f"AI-DRIVEN WORKFORCE INTELLIGENCE REPORT - {report_type.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        
        # Metadata Block
        ws['A3'] = "Generated Date:"
        ws['A3'].font = bold_font
        ws['B3'] = datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M:%S UTC")
        ws['B3'].font = regular_font
        
        ws['A4'] = "Requested By:"
        ws['A4'].font = bold_font
        ws['B4'] = f"{current_user.name} ({current_user.role.value})"
        ws['B4'].font = regular_font

        # AI Insights Summary Block
        ws.merge_cells('A6:E6')
        ws['A6'] = "AI EXECUTIVE STRATEGIC INSIGHTS"
        ws['A6'].font = header_font
        ws['A6'].fill = header_fill
        ws['A6'].alignment = Alignment(horizontal='left', indent=1)
        ws.row_dimensions[6].height = 20

        # Build dynamic summary content
        summary_lines = []
        if report_type == "productivity" or report_type == "executive":
            summary_lines.append(f"• Workforce operational productivity index averages {perf_intel['team_average_productivity']}%.")
            summary_lines.append(f"• Active workload contains {task_intel['total_active_tasks']} tasks under tracking.")
            if task_intel['overloaded_employees']:
                summary_lines.append(f"• Workload capacity warning: {len(task_intel['overloaded_employees'])} employees exceed active thresholds.")
        if report_type == "payroll" or report_type == "executive":
            summary_lines.append(f"• AI payroll scan flagged {len(payroll_intel.get('alerts', []))} variance anomaly alerts.")
        if report_type == "attendance" or report_type == "executive":
            summary_lines.append(f"• Late login check-ins flagged {len(attendance_intel.get('late_login_trends', []))} persistent trends in current 30-day window.")
        
        ws.merge_cells('A7:E9')
        ws['A7'] = "\n".join(summary_lines)
        ws['A7'].font = Font(name='Arial', size=10, italic=True)
        ws['A7'].alignment = Alignment(vertical='top', wrap_text=True)
        ws['A7'].fill = accent_fill
        ws.row_dimensions[7].height = 50

        # Alert warnings
        ws.merge_cells('A11:E11')
        ws['A11'] = "CRITICAL OPERATIONAL WARNINGS FLAGGED BY AI"
        ws['A11'].font = header_font
        ws['A11'].fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
        ws['A11'].alignment = Alignment(horizontal='left', indent=1)
        ws.row_dimensions[11].height = 20

        alerts_list = []
        alerts_list.extend(task_intel.get("overloaded_employees", []))
        alerts_list.extend(attendance_intel.get("alerts", []))
        alerts_list.extend(payroll_intel.get("alerts", []))
        
        ws.merge_cells('A12:E14')
        if alerts_list:
            ws['A12'] = "\n".join([f"⚠ {str(a['details']) if isinstance(a, dict) else str(a)}" for a in alerts_list[:3]])
        else:
            ws['A12'] = "✔ Operational status normal. Zero anomalies flagged in this cycle."
        ws['A12'].font = Font(name='Arial', size=10, color='721C24' if alerts_list else '155724')
        ws['A12'].alignment = Alignment(vertical='top', wrap_text=True)
        ws['A12'].fill = alert_fill if alerts_list else PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        ws.row_dimensions[12].height = 50

        # Report Detailed Data Table
        ws['A16'] = "DETAILED ANALYTICS DATA"
        ws['A16'].font = bold_font
        
        start_row = 18
        if report_type == "productivity":
            headers = ["EMPLOYEE NAME", "ROLE", "TASKS ASSIGNED", "TASKS COMPLETED", "PRODUCTIVITY SCORE (%)", "BURNOUT RISK"]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            curr_row = start_row + 1
            for p in perf_intel["employee_performance"]:
                ws.cell(row=curr_row, column=1, value=p["name"]).font = regular_font
                ws.cell(row=curr_row, column=2, value=p["role"].upper()).font = regular_font
                ws.cell(row=curr_row, column=3, value=p["tasks_assigned"]).font = regular_font
                ws.cell(row=curr_row, column=4, value=p["tasks_completed"]).font = regular_font
                ws.cell(row=curr_row, column=5, value=p["productivity_score"]).font = regular_font
                ws.cell(row=curr_row, column=6, value=p["burnout_risk"]).font = Font(name='Arial', size=10, bold=p["burnout_risk"]=="High", color='D32F2F' if p["burnout_risk"]=="High" else '000000')
                
                # Apply borders
                for c in range(1, 7):
                    ws.cell(row=curr_row, column=c).border = thin_border
                curr_row += 1
                
        elif report_type == "attendance":
            headers = ["EMPLOYEE NAME", "TOTAL LOGS (30D)", "CONSISTENCY SCORE (%)", "LATE TRENDS WARNING", "ABSENTEE WARNING"]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            curr_row = start_row + 1
            for c in attendance_intel["consistency_rankings"]:
                # Check warnings
                late_warn = next((l["details"] for l in attendance_intel["late_login_trends"] if l["user_id"] == c["user_id"]), "Normal")
                abs_warn = next((a["details"] for a in attendance_intel["absentee_warnings"] if a["user_id"] == c["user_id"]), "Normal")
                
                ws.cell(row=curr_row, column=1, value=c["name"]).font = regular_font
                ws.cell(row=curr_row, column=2, value=c["total_logs"]).font = regular_font
                ws.cell(row=curr_row, column=3, value=c["consistency_score"]).font = regular_font
                ws.cell(row=curr_row, column=4, value=late_warn).font = regular_font
                ws.cell(row=curr_row, column=5, value=abs_warn).font = regular_font
                
                for col in range(1, 6):
                    ws.cell(row=curr_row, column=col).border = thin_border
                curr_row += 1
                
        else: # Default Task/Executive general data
            headers = ["TASK DESCRIPTION", "ASSIGNEE", "PRIORITY", "DUE DATE", "COMPLETION RISK", "DELAY PROBABILITY (%)"]
            for col_idx, text in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            curr_row = start_row + 1
            for t in task_intel["task_predictions"][:15]:
                ws.cell(row=curr_row, column=1, value=t["description"]).font = regular_font
                ws.cell(row=curr_row, column=2, value=t["assigned_to_name"]).font = regular_font
                ws.cell(row=curr_row, column=3, value=t["priority"].upper()).font = regular_font
                ws.cell(row=curr_row, column=4, value=t["deadline"]).font = regular_font
                ws.cell(row=curr_row, column=5, value=t["completion_prediction"]).font = Font(name='Arial', size=10, bold=t["completion_prediction"]!="On Time", color='D32F2F' if t["completion_prediction"]!="On Time" else '2E7D32')
                ws.cell(row=curr_row, column=6, value=t["risk_score"]).font = regular_font
                
                for col in range(1, 7):
                    ws.cell(row=curr_row, column=col).border = thin_border
                curr_row += 1

        # Adjust columns width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=ai_{report_type}_report.xlsx"},
        )
        
    else:
        # 3. HTML Print-ready template (replaces PDF compilation perfectly and safely)
        summary_bullets = "".join([
            f"<li style='margin-bottom: 8px;'><b>Productivity Index:</b> Scoped team productivity averages {perf_intel['team_average_productivity']}%.</li>",
            f"<li style='margin-bottom: 8px;'><b>Task Capacity:</b> {task_intel['total_active_tasks']} active tracking items, with {task_intel['total_overdue_tasks']} currently overdue.</li>",
            f"<li style='margin-bottom: 8px;'><b>Burnout Indicators:</b> Flagged {len([p for p in perf_intel['employee_performance'] if p['burnout_risk']=='High'])} personnel under high stress alerts.</li>"
        ])
        
        alerts_li = ""
        alerts_list = []
        alerts_list.extend(task_intel.get("overloaded_employees", []))
        alerts_list.extend(attendance_intel.get("alerts", []))
        alerts_list.extend(payroll_intel.get("alerts", []))
        
        if alerts_list:
            alerts_li = "".join([f"<li style='color: #D32F2F; margin-bottom: 6px;'>⚠ {str(a['details']) if isinstance(a, dict) else str(a)}</li>" for a in alerts_list[:4]])
        else:
            alerts_li = "<li style='color: #2E7D32;'>✔ Operations Normal: No anomalies detected in current audit.</li>"

        rows_html = ""
        if report_type == "productivity":
            rows_html = "".join([
                f"<tr>"
                f"<td>{p['name']}</td>"
                f"<td>{p['role'].replace('_', ' ').title()}</td>"
                f"<td>{p['tasks_assigned']}</td>"
                f"<td>{p['tasks_completed']}</td>"
                f"<td><span style='font-weight: bold;'>{p['productivity_score']}%</span></td>"
                f"<td style='color: {'#D32F2F' if p['burnout_risk']=='High' else '#000000'}; font-weight: { 'bold' if p['burnout_risk']=='High' else 'normal' };'>{p['burnout_risk']}</td>"
                f"</tr>"
                for p in perf_intel["employee_performance"]
            ])
            table_headers = "<th>Employee Name</th><th>Designation</th><th>Assigned</th><th>Completed</th><th>Productivity Score</th><th>Burnout Risk</th>"
        elif report_type == "attendance":
            table_headers = "<th>Employee Name</th><th>Total Check-ins (30D)</th><th>Consistency Score</th><th>Late Warnings</th><th>Absentee Warnings</th>"
            rows_html = ""
            for c in attendance_intel["consistency_rankings"]:
                late_warn = next((l["details"] for l in attendance_intel["late_login_trends"] if l["user_id"] == c["user_id"]), "Normal")
                abs_warn = next((a["details"] for a in attendance_intel["absentee_warnings"] if a["user_id"] == c["user_id"]), "Normal")
                rows_html += (
                    f"<tr>"
                    f"<td>{c['name']}</td>"
                    f"<td>{c['total_logs']} logs</td>"
                    f"<td><span style='font-weight: bold;'>{c['consistency_score']}%</span></td>"
                    f"<td>{late_warn}</td>"
                    f"<td>{abs_warn}</td>"
                    f"</tr>"
                )
        else:
            table_headers = "<th>Task Description</th><th>Assignee</th><th>Priority</th><th>Due Date</th><th>Completion prediction</th><th>Risk Score</th>"
            rows_html = "".join([
                f"<tr>"
                f"<td>{t['description']}</td>"
                f"<td>{t['assigned_to_name']}</td>"
                f"<td style='text-transform: uppercase;'>{t['priority']}</td>"
                f"<td>{t['deadline'][:16].replace('T', ' ')}</td>"
                f"<td style='color: {'#D32F2F' if t['completion_prediction']!='On Time' else '#2E7D32'}; font-weight: bold;'>{t['completion_prediction']}</td>"
                f"<td>{t['risk_score']}%</td>"
                f"</tr>"
                for t in task_intel["task_predictions"][:15]
            ])

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>TaskReward AI Report - {report_type.title()}</title>
            <style>
                body {{
                    font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif;
                    color: #333;
                    margin: 0;
                    padding: 30px;
                    background: #FFF;
                }}
                .header {{
                    display: flex;
                    justify-content: space-between;
                    border-bottom: 2px solid #3F51B5;
                    padding-bottom: 15px;
                    margin-bottom: 30px;
                }}
                .brand {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #3F51B5;
                }}
                .subtitle {{
                    font-size: 11px;
                    text-transform: uppercase;
                    color: #777;
                    letter-spacing: 1.5px;
                    margin-top: 5px;
                }}
                .report-title {{
                    font-size: 20px;
                    font-weight: 900;
                    color: #212121;
                    margin: 0 0 20px 0;
                }}
                .meta-table {{
                    font-size: 12px;
                    margin-bottom: 25px;
                }}
                .meta-table td {{
                    padding: 4px 10px 4px 0;
                }}
                .insight-box {{
                    background: #F5F6FD;
                    border-left: 5px solid #3F51B5;
                    padding: 15px 20px;
                    border-radius: 4px;
                    margin-bottom: 25px;
                }}
                .insight-box h3 {{
                    margin: 0 0 10px 0;
                    font-size: 14px;
                    color: #3F51B5;
                    text-transform: uppercase;
                    letter-spacing: 1px;
                }}
                .alert-box {{
                    background: #FFF5F5;
                    border-left: 5px solid #E53935;
                    padding: 15px 20px;
                    border-radius: 4px;
                    margin-bottom: 35px;
                }}
                .alert-box h3 {{
                    margin: 0 0 10px 0;
                    font-size: 14px;
                    color: #E53935;
                    text-transform: uppercase;
                    letter-spacing: 1px;
                }}
                table.data-table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-top: 15px;
                    font-size: 12px;
                }}
                table.data-table th {{
                    background: #3F51B5;
                    color: #FFF;
                    text-align: left;
                    padding: 10px 12px;
                    font-weight: bold;
                }}
                table.data-table td {{
                    padding: 10px 12px;
                    border-bottom: 1px solid #E0E0E0;
                }}
                table.data-table tr:nth-child(even) {{
                    background: #FAFAFA;
                }}
                .footer {{
                    margin-top: 50px;
                    border-top: 1px solid #E0E0E0;
                    padding-top: 15px;
                    font-size: 10px;
                    color: #777;
                    text-align: center;
                }}
                @media print {{
                    body {{ padding: 0; }}
                    button.print-btn {{ display: none; }}
                }}
                button.print-btn {{
                    background: #3F51B5;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    font-size: 14px;
                    font-weight: bold;
                    border-radius: 8px;
                    cursor: pointer;
                    float: right;
                    transition: opacity 0.2s;
                }}
                button.print-btn:hover {{
                    opacity: 0.9;
                }}
            </style>
        </head>
        <body>
            <button class="print-btn" onclick="window.print()">Print to PDF</button>
            <div class="header">
                <div>
                    <div class="brand">TaskReward</div>
                    <div class="subtitle">AI Workforce Management Solutions</div>
                </div>
                <div style="text-align: right; font-size: 12px; color: #555;">
                    <div>Confidential Document</div>
                    <div>Operational Intelligence Unit</div>
                </div>
            </div>

            <h1 class="report-title">{report_type.replace('_', ' ').title()} Performance Audit</h1>
            
            <table class="meta-table">
                <tr>
                    <td><b>Generated Date:</b></td>
                    <td>{datetime.now(timezone.utc).strftime("%d-%b-%Y %H:%M:%S UTC")}</td>
                </tr>
                <tr>
                    <td><b>Report Scope:</b></td>
                    <td>{report_type.replace('_', ' ').upper()} ANALYTICS</td>
                </tr>
                <tr>
                    <td><b>Requested By:</b></td>
                    <td>{current_user.name} ({current_user.role.value.upper()})</td>
                </tr>
            </table>

            <div class="insight-box">
                <h3>AI Strategic Summary</h3>
                <ul style="margin: 0; padding-left: 20px; font-size: 13px; line-height: 1.5;">
                    {summary_bullets}
                </ul>
            </div>

            <div class="alert-box">
                <h3>Anomaly Scanner Alerts</h3>
                <ul style="margin: 0; padding-left: 20px; font-size: 13px; line-height: 1.5;">
                    {alerts_li}
                </ul>
            </div>

            <h2 style="font-size: 15px; margin-top: 30px; border-bottom: 1px solid #CCC; padding-bottom: 5px;">Detailed Scoped Records</h2>
            <table class="data-table">
                <thead>
                    <tr>
                        {table_headers}
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>

            <div class="footer">
                This document is generated by TaskReward AI. Content is synthesized dynamically based on real-time workforce databases.
            </div>
        </body>
        </html>
        """
        return HTMLResponse(content=html_content)
